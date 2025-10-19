import pandas as pd
from datetime import datetime, timedelta
import psycopg2
from psycopg2 import sql
import statistics
from datetime import datetime
from math import exp
import math
from openpyxl import Workbook

# --- PostgreSQL connection (adjust creds/host/port/db as needed) ---
conn = psycopg2.connect(
    host="",
    port="",
    database="",
    user="",
    password=""
)
cursor = conn.cursor()

# --- Input window & target device ---
start_date = datetime(2022, 7, 1)
end_date   = datetime(2023, 9, 30)
device_id  = 789

# --- Output collectors (per-day ET0) ---
output_date_list: list[str] = []
output_list: list[str]      = []

# --- Iterate day-by-day over [start_date, end_date] ---
current_date = start_date
while current_date <= end_date:
    list_date = f"{current_date.day:02d}/{current_date.month:02d}/{current_date.year}"
    year  = current_date.year
    month = current_date.month
    day   = current_date.day

    # ISO date string for DB function
    date_string = datetime(year, month, day).strftime("%Y-%m-%d")

    # --- Call utils.fn_get_et0_data(device_id, start, end) ---
    query = sql.SQL("SELECT * FROM utils.fn_get_et0_data(%s, %s, %s)")
    arguments = (device_id, date_string, date_string)
    full_query = query.as_string(conn)
    cursor.execute(full_query, arguments)
    results = cursor.fetchall()

    # --- Extract hourly weather series (°C, %, m/s, W/m²) & site meta ---
    try:
        temperature_list = [float(row[8]) for row in results]
        humidity_list    = [float(row[6]) for row in results]
        wind_list        = [float(row[5]) for row in results]
        radiation_list   = [float(row[7]) for row in results]
        wind_sensor_height = float(results[0][1])  # m (unused below)
        latitude           = float(results[0][2])  # degrees in DMS-like format
    except Exception:
        current_date += timedelta(days=1)
        continue  # skip day if data is incomplete

    # --- Fixed site constants (tune for your station) ---
    altitude = 8  # m

    # --- Daily aggregates (T, RH, wind, shortwave) ---
    try:
        daily_max_temperature = max(temperature_list)
        daily_min_temperature = min(temperature_list)
        daily_mean_temperature = (daily_max_temperature + daily_min_temperature) / 2

        daily_max_relative_humidity = max(humidity_list)
        daily_min_relative_humidity = min(humidity_list)

        # wind_list likely in km/h → convert to m/s
        wind_speed = statistics.mean(wind_list) * 1000 / 3600

        # radiation_list likely in W/m² → MJ/m²/day factor 0.0864 for daily mean
        solar_radiation = statistics.mean(radiation_list) * 0.0864
    except Exception:
        current_date += timedelta(days=1)
        continue

    # --- FAO-56 constants ---
    stefan_boltzmann_constant   = 4.903 * (10 ** -9)
    latent_heat_of_vaporization = 2.45
    solar_constant              = 0.0820
    albedo                      = 0.23

    # --- Psychrometrics & vapor pressure terms ---
    saturation_vapour_pressure = 0.6108 * exp((17.27 * daily_mean_temperature) / (daily_mean_temperature + 237.3))
    atmospheric_pressure = 101.3 * ((293 - 0.0065 * altitude) / 293) ** 5.26
    slope_of_svp = 4098 * saturation_vapour_pressure / ((daily_mean_temperature + 237.3) ** 2)  # kPa/°C
    psychrometric_constant = 0.000665 * atmospheric_pressure

    # --- Combination equation weights (FAO-56 Penman–Monteith) ---
    var1 = 1 + 0.34 * wind_speed
    var2 = slope_of_svp / (slope_of_svp + psychrometric_constant * var1)
    var3 = psychrometric_constant / (slope_of_svp + psychrometric_constant * var1)
    var4 = 900 / (daily_mean_temperature + 273) * wind_speed

    # --- Saturation vapor pressure at Tmax/Tmin (kPa) ---
    es_Tmax = 0.6108 * exp(17.27 * daily_max_temperature / (daily_max_temperature + 237.3))
    es_Tmin = 0.6108 * exp(17.27 * daily_min_temperature / (daily_min_temperature + 237.3))
    es_mean = (es_Tmin + es_Tmax) / 2

    # --- Actual vapor pressure (ea) and deficit (VPD) ---
    ea = (es_Tmin * (daily_max_relative_humidity / 100) + es_Tmax * (daily_min_relative_humidity / 100)) / 2
    vpd = es_mean - ea

    # --- Day of year (Julian-like) for extraterrestrial radiation ---
    J = int(275 * month / 9 - 30 + day) - 2
    if month < 3:
        J += 2
    elif month > 2 and year % 4 == 0:
        J += 1

    # --- Latitude in decimal degrees (input appears as DDMM.m-like) ---
    decimal_part     = int(latitude)
    fractional_part  = latitude - decimal_part
    decimal_degrees_latitude = decimal_part + (fractional_part / 60 * 100)
    radians_latitude = math.pi / 180 * decimal_degrees_latitude

    # --- Extraterrestrial radiation Ra (MJ/m²/day) ---
    dr  = 1 + 0.033 * math.cos((2 * math.pi * J) / 365)                         # inverse relative distance
    sol_dec = 0.409 * math.sin((2 * math.pi * J / 365) - 1.39)                  # solar declination
    ws  = math.acos(-1 * math.tan(radians_latitude) * math.tan(sol_dec))        # sunset hour angle
    Ra  = (24 * 60 / math.pi) * solar_constant * dr * (
          ws * math.sin(radians_latitude) * math.sin(sol_dec)
          + math.cos(radians_latitude) * math.cos(sol_dec) * math.sin(ws))

    # --- Clear-sky shortwave & net radiation (MJ/m²/day) ---
    N     = 24 * ws / math.pi                                  # daylight hours (unused)
    Rso   = (0.75 + 2e-5 * altitude) * Ra                       # clear-sky radiation
    Rs_Rso = solar_radiation / Rso
    Rs_Rso = 1 if Rs_Rso > 1 else (0.3 if Rs_Rso < 0.3 else Rs_Rso)

    Rns = (1 - albedo) * solar_radiation                        # net shortwave
    Rnl = (stefan_boltzmann_constant *
           ((((daily_max_temperature + 273.6) ** 4) + ((daily_min_temperature + 273.6) ** 4)) / 2) *
           (0.34 - 0.14 * math.sqrt(ea)) *
           (1.35 * Rs_Rso - 0.35))                               # net longwave

    Rn = Rns - Rnl                                               # net radiation
    G  = 0                                                       # soil heat flux ≈ 0 for daily

    # --- FAO-56 reference ET0 (mm/day) ---
    term_rad  = 0.408 * (Rn - G) * var2
    term_aero = var4 * vpd * var3
    et0 = term_rad + term_aero

    # --- Collect result for this day ---
    output_date_list.append(list_date)
    output_list.append(f"{et0}")

    current_date += timedelta(days=1)

# --- Write results to Excel (date | et0) ---
excel_list: dict[str, list[str]] = {k: [v] for k, v in zip(output_date_list, output_list)}

workbook = Workbook()
sheet = workbook.active

row = 1
for key, value in excel_list.items():
    sheet.cell(row=row, column=1, value=key)  # date
    for i, val in enumerate(value, start=1):
        sheet.cell(row=row + i - 1, column=2, value=val)  # et0
    row += len(value)

workbook.save("Results/ciftlik_3_2022_daily.xlsx")
print("İşlem tamamlandı.")
